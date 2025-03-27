import os
import time
import openpyxl
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from requests.exceptions import ConnectionError
from urllib3.exceptions import NewConnectionError
from selenium.common.exceptions import WebDriverException

def initialize_driver(headless=True):
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--allow-running-insecure-content')
    #if headless:
    #   options.add_argument('--headless')
    driver = webdriver.Chrome(options=options)
    return driver

def login(driver, user_id_str, password_str):
    driver.get("website_link")
    driver.find_element(By.XPATH, '/html/body/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[6]/td/table/tbody/tr[2]/td[3]/font/input').send_keys(user_id_str)
    driver.find_element(By.XPATH, '/html/body/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[6]/td/table/tbody/tr[3]/td[3]/font/input').send_keys(password_str)
    driver.find_element(By.XPATH, '/html/body/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[6]/td/table/tbody/tr[4]/td[2]/font/input[1]').click()
    time.sleep(2)

def navigate_to_projects(driver):
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/table/tbody/tr[3]/td/input[5]'))).click()
        time.sleep(1)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/table[3]/tbody/tr/td[1]/table/tbody/tr[2]/td/a/b'))).click()
        time.sleep(1)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/table[3]/tbody/tr/td[2]/table/tbody/tr[2]/td/table/tbody/tr/td/blockquote/ul/li[1]/a'))).click()
        time.sleep(1)
    except Exception as e:
        print(f"An error occurred during navigation: {e}")
        driver.quit()
        exit()

def search_projects(driver):
    try:
        # Select year from dropdown menu
        year_dropdown = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/form[1]/table/tbody/tr[4]/td[2]/select')))
        select = Select(year_dropdown)
        select.select_by_visible_text("2019 / 20")
        time.sleep(1)
        
        # Click the search button
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form[1]/table/tbody/tr[7]/td/input'))).click()
        time.sleep(1)
    except Exception as e:
        print(f"An error occurred while selecting the year and clicking the search button: {e}")
        driver.quit()
        exit()

def extract_project_data(driver, row_index):
    try:
        project_number = driver.find_element(By.XPATH, f'/html/body/table[5]/tbody/tr[{row_index}]/td[1]/input[1]').get_attribute('value').strip()
        title = driver.find_element(By.XPATH, f'/html/body/table[5]/tbody/tr[{row_index}]/td[2]').text.strip()
        pi_name = driver.find_element(By.XPATH, f'/html/body/table[5]/tbody/tr[{row_index}]/td[3]').text.strip()
        result = driver.find_element(By.XPATH, f'/html/body/table[5]/tbody/tr[{row_index}]/td[4]').text.strip()
        detail_button = driver.find_element(By.XPATH, f'/html/body/table[5]/tbody/tr[{row_index}]/td[1]/input[1]')
        time.sleep(1)
        
        return {
            'Project Number': project_number,
            'Title': title,
            'PI Name': pi_name,
            'Result': result,
            'Detail Button': detail_button
        }
    except Exception as e:
        print(f"An error occurred while extracting data from the main page: {e}")
        raise

def extract_project_detail(driver, project):
    try:
        project['Detail Button'].click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[-1])
        funding_scheme = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/table[4]/tbody/tr[1]/td[2]'))).text.strip()
        project['Funding Scheme'] = funding_scheme
        
        if funding_scheme == "Early Career Scheme (ECS)":
            project['Project Number'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[2]/td[2]').text.strip()
            project['Exercise Year'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[3]/td[2]').text.strip()
            project['Title'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[4]/td[2]').text.strip()
            project['PI Name'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[5]/td[2]').text.strip()
            project['Institution'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[6]/td[2]').text.strip()
            project['Project Fund'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[7]/td[2]').text.strip()
            project['ECS Award'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[8]/td[2]').text.strip()
            project['Total Fund Awarded'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[9]/td[2]').text.strip()
            project['Result Score'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[10]/td[2]').text.strip()
            project['Approved Project duration'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[11]/td[2]').text.strip()
            project['Notes for the Applicants'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[12]/td[2]/u/a').text.strip()
        else:
            project['ECS Award'] = "NA"
            project['Total Fund Awarded'] = "NA"
            project['Project Number'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[2]/td[2]').text.strip()
            project['Exercise Year'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[3]/td[2]').text.strip()
            project['Institution'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[6]/td[2]').text.strip()
            project['Project Fund'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[7]/td[2]').text.strip()
            project['Result Score'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[8]/td[2]').text.strip()
            project['Approved Project duration'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[9]/td[2]').text.strip()
            project['Notes for the Applicants'] = driver.find_element(By.XPATH, '/html/body/table[4]/tbody/tr[10]/td[2]').text.strip()
    except Exception as e:
        print(f"An error occurred while extracting data from the detail page: {e}")

def navigate_comments_from_panel(driver, project):
    try:
        # Click the "View Comments" button to open the third tab
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/input[1]'))).click()
        time.sleep(1)
        
        # Switch to the third tab
        driver.switch_to.window(driver.window_handles[-1])
        
        # Click the "Comments from Panel" button
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/p[1]/input'))).click()
        time.sleep(1)
        
        # Scrape the comments from the panel
        comments_elements = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, '/html/body/table[5]/tbody/tr/td/table/tbody/tr/td'))
        )
        comments = " ".join([comment.text.strip() for comment in comments_elements])
        project['Comments from Panel'] = comments
        
        # Click the "Return" button to go back to the second tab
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/input[4]'))).click()
        time.sleep(1)
        
    except Exception as e:
        print(f"An error occurred while navigating comments from panel: {e}")

def navigate_external_reviewer_comments(driver, project):
    try:
        # Scrape external reviewer comments
        x = 2
        external_comments = []
        while True:
            try:
                review_button_xpath = f'/html/body/form/p[2]/table[2]/tbody/tr[{x}]/td[2]/input'
                review_button = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, review_button_xpath)))
                review_button.click()
                time.sleep(1)
                
                # Switch to the new window if a new window is opened
                driver.switch_to.window(driver.window_handles[-1])
                
                # Perform actions on the detailed comments page
                # Scrape the comments and checkbox for each section
                def get_comment(xpath):
                    try:
                        return driver.find_element(By.XPATH, xpath).text.strip()
                    except:
                        return ""
                
                def get_checkbox_value(radio_xpaths):
                    for xpath in radio_xpaths:
                        try:
                            element = driver.find_element(By.XPATH, xpath)
                            if element.is_selected():
                                return element.get_attribute('value')
                        except:
                            continue
                    return ""

                def get_checkbox_label(value, section):
                    labels = {
                        "1": "Excellent",
                        "2": "Very Good",
                        "3": "Good",
                        "4": "Fair",
                        "5": "Poor"
                    }
                    if section == "duration":
                        labels = {
                            "1": "Too Long",
                            "2": "Appropriate",
                            "3": "Too Short"
                        }
                    elif section == "impact":
                        labels = {
                            "1": "High",
                            "2": "Moderate",
                            "3": "Low",
                            "4": "None"
                        }
                    return labels.get(value, "")

                scientific_merit = get_checkbox_label(get_checkbox_value([
                    '/html/body/form/p[3]/table[1]/tbody/tr[2]/td[1]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[2]/td[2]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[2]/td[3]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[2]/td[4]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[2]/td[5]/input'
                ]), "default")

                duration_proposed = get_checkbox_label(get_checkbox_value([
                    '/html/body/form/p[3]/table[1]/tbody/tr[4]/td[1]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[4]/td[2]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[4]/td[3]/input'
                ]), "duration")

                impact_of_research = get_checkbox_label(get_checkbox_value([
                    '/html/body/form/p[3]/table[1]/tbody/tr[6]/td[1]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[6]/td[2]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[6]/td[3]/input',
                    '/html/body/form/p[3]/table[1]/tbody/tr[6]/td[4]/input'
                ]), "impact")

                ability_to_undertake = get_checkbox_label(get_checkbox_value([
                    '/html/body/form/p[3]/table[2]/tbody/tr[2]/td[1]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[2]/td[2]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[2]/td[3]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[2]/td[4]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[2]/td[5]/input'
                ]), "default")

                track_record = get_checkbox_label(get_checkbox_value([
                    '/html/body/form/p[3]/table[2]/tbody/tr[4]/td[1]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[4]/td[2]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[4]/td[3]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[4]/td[4]/input',
                    '/html/body/form/p[3]/table[2]/tbody/tr[4]/td[5]/input'
                ]), "default")

                seq_no_comments = {
                    'Seq No': x-1,
                    'Objective Evaluation': get_checkbox_label(get_checkbox_value([
                        '/html/body/form/table[2]/tbody/tr[2]/td[1]/input',
                        '/html/body/form/table[2]/tbody/tr[2]/td[2]/input',
                        '/html/body/form/table[2]/tbody/tr[2]/td[3]/input',
                        '/html/body/form/table[2]/tbody/tr[2]/td[4]/input',
                        '/html/body/form/table[2]/tbody/tr[2]/td[5]/input'
                    ]), "default"),
                    'Objective Comments': get_comment('/html/body/form/table[3]/tbody/tr[2]/td/table/tbody/tr/td'),
                    'Research Design Evaluation': get_checkbox_label(get_checkbox_value([
                        '/html/body/form/table[4]/tbody/tr[2]/td[1]/input',
                        '/html/body/form/table[4]/tbody/tr[2]/td[2]/input',
                        '/html/body/form/table[4]/tbody/tr[2]/td[3]/input',
                        '/html/body/form/table[4]/tbody/tr[2]/td[4]/input',
                        '/html/body/form/table[4]/tbody/tr[2]/td[5]/input'
                    ]), "default"),
                    'Research Design Comments': get_comment('/html/body/form/table[5]/tbody/tr[2]/td/table/tbody/tr/td'),
                    'Feasibility Evaluation': get_checkbox_label(get_checkbox_value([
                        '/html/body/form/table[6]/tbody/tr[2]/td[1]/input',
                        '/html/body/form/table[6]/tbody/tr[2]/td[2]/input',
                        '/html/body/form/table[6]/tbody/tr[2]/td[3]/input',
                        '/html/body/form/table[6]/tbody/tr[2]/td[4]/input',
                        '/html/body/form/table[6]/tbody/tr[2]/td[5]/input'
                    ]), "default"),
                    'Feasibility Comments': get_comment('/html/body/form/table[7]/tbody/tr[2]/td/table/tbody/tr/td'),
                    'Most Original Aspect Comments': get_comment('/html/body/form/table[8]/tbody/tr[2]/td/table/tbody/tr/td'),
                    'Budget and Planning Comments': get_comment('/html/body/form/table[9]/tbody/tr[2]/td/table/tbody/tr/td'),
                    'Overall Comments': get_comment('/html/body/form/table[10]/tbody/tr/td/table/tbody/tr/td'),
                    'Scientific Merit': scientific_merit,
                    'Duration Proposed': duration_proposed,
                    'Impact of Research': impact_of_research,
                    'Ability to Undertake Proposal': ability_to_undertake,
                    'Track Record in Field': track_record
                }
                external_comments.append(seq_no_comments)
                
                # Click the "Return" button to go back to the "Seq No." list page
                return_button_xpath = '/html/body/form/p[4]/table/tbody/tr/td/input'
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, return_button_xpath))).click()
                time.sleep(1)
                  
                x += 1
            except Exception as e:
                print(f"No more external reviewer comments found after {x-2} reviews. Stopping.")
                break
        
        project['External Reviewer Comments'] = external_comments
        
    except Exception as e:
        print(f"An error occurred while navigating external reviewer comments: {e}")

def close_tabs_and_return(driver):
    try:
        # Close the 3rd tab by clicking the close button
        driver.switch_to.window(driver.window_handles[-1])
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/p[3]/input'))).click()
        time.sleep(1)
        
        # Close the 2nd tab by clicking the close button
        driver.switch_to.window(driver.window_handles[-1])
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/input[2]'))).click()
        time.sleep(1)
        
        # Switch back to the main window
        driver.switch_to.window(driver.window_handles[0])
        
    except Exception as e:
        print(f"An error occurred while closing tabs and returning: {e}")

def clean_value(value):
    if isinstance(value, str):
        # Remove or replace characters that cannot be used in Excel
        return ''.join(c for c in value if ord(c) >= 32 and ord(c) != 127)
    return value

def save_to_excel(projects, filename='project_data_2019.xlsx'):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws1 = wb["Project Data"]
        ws2 = wb["External Reviewer Comments"]
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Project Data"
        headers1 = [
            'Project Number', 'Title', 'PI Name', 'Result', 'Funding Scheme',
            'Exercise Year', 'Institution', 'Project Fund', 'ECS Award', 'Total Fund Awarded', 'Result Score',
            'Approved Project duration', 'Notes for the Applicants', 'Comments from Panel'
        ]
        ws1.append(headers1)
        
        ws2 = wb.create_sheet(title="External Reviewer Comments")
        headers2 = [
            'Project Number', 'Title', 'PI Name', 'External Reviewer No', 'Objective Evaluation', 'Objective Comments',
            'Research Design Evaluation', 'Research Design Comments', 'Feasibility Evaluation', 'Feasibility Comments',
            'Most Original Aspect Comments', 'Budget and Planning Comments', 'Overall Comments',
            'Scientific Merit', 'Duration Proposed', 'Impact of Research',
            'Ability to Undertake Proposal', 'Track Record in Field'
        ]
        ws2.append(headers2)
    
    for project in projects:
        row1 = [
            clean_value(project.get('Project Number', '')),
            clean_value(project.get('Title', '')),
            clean_value(project.get('PI Name', '')),
            clean_value(project.get('Result', '')),
            clean_value(project.get('Funding Scheme', '')),
            clean_value(project.get('Exercise Year', '')),
            clean_value(project.get('Institution', '')),
            clean_value(project.get('Project Fund', '')),
            clean_value(project.get('ECS Award', 'NA')),
            clean_value(project.get('Total Fund Awarded', 'NA')),
            clean_value(project.get('Result Score', '')),
            clean_value(project.get('Approved Project duration', '')),
            clean_value(project.get('Notes for the Applicants', '')),
            clean_value(project.get('Comments from Panel', ''))
        ]
        ws1.append(row1)
        
        for seq_no_comments in project.get('External Reviewer Comments', []):
            row2 = [
                clean_value(project.get('Project Number', '')),
                clean_value(project.get('Title', '')),
                clean_value(project.get('PI Name', '')),
                clean_value(seq_no_comments.get('Seq No', '')),
                clean_value(seq_no_comments.get('Objective Evaluation', '')),
                clean_value(seq_no_comments.get('Objective Comments', '')),
                clean_value(seq_no_comments.get('Research Design Evaluation', '')),
                clean_value(seq_no_comments.get('Research Design Comments', '')),
                clean_value(seq_no_comments.get('Feasibility Evaluation', '')),
                clean_value(seq_no_comments.get('Feasibility Comments', '')),
                clean_value(seq_no_comments.get('Most Original Aspect Comments', '')),
                clean_value(seq_no_comments.get('Budget and Planning Comments', '')),
                clean_value(seq_no_comments.get('Overall Comments', '')),
                clean_value(seq_no_comments.get('Scientific Merit', '')),
                clean_value(seq_no_comments.get('Duration Proposed', '')),
                clean_value(seq_no_comments.get('Impact of Research', '')),
                clean_value(seq_no_comments.get('Ability to Undertake Proposal', '')),
                clean_value(seq_no_comments.get('Track Record in Field', ''))
            ]
            ws2.append(row2)
    
    wb.save(filename)
    print(f"Data saved to {filename}")

def main(start_row_index=2):
    driver = initialize_driver()
    login(driver, "login", "password")
    navigate_to_projects(driver)
    search_projects(driver)
    row_index = start_row_index
    projects = []
    while True:
        try:
            project = extract_project_data(driver, row_index)
            extract_project_detail(driver, project)
            navigate_comments_from_panel(driver, project)
            navigate_external_reviewer_comments(driver, project)
            close_tabs_and_return(driver)
            projects.append(project)
            row_index += 1

            # Save progress after each project
            save_to_excel(projects)
            projects = []  # Reset projects list to avoid duplicate saving

        except WebDriverException as e:
            print(f"An error occurred while processing project at row {row_index}: {e}")
            # Handle browser crash by restarting and continuing
            driver.quit()
            driver = initialize_driver()
            login(driver, "slim16", "RO2230cyc#")
            navigate_to_projects(driver)
            search_projects(driver)
            time.sleep(3)
            # Continue from the same project
            continue

        except Exception as e:
            print(f"An error occurred while processing project at row {row_index}: {e}")
            save_to_excel(projects)
            break
    driver.quit()
    save_to_excel(projects)

if __name__ == "__main__":
    start_row = int(input("Enter the starting project row index: "))
    main(start_row)